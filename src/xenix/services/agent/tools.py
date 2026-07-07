from __future__ import annotations

import csv
import time
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl
from pydantic import BaseModel, Field, ValidationError as PydanticValidationError

from ...config import AppPaths
from ...exceptions import NotFoundError, ValidationError
from ..analysis_graph import AnalysisGraphService, GraphDatasetInput
from ..analysis_lambda import AnalysisLambdaDataset, AnalysisLambdaInput, AnalysisLambdaService
from ..analysis_profile import AnalysisProfileService, ProfileDatasetInput
from ..artifact_service import (
    ArtifactService,
    RegisterArtifactInput,
    build_artifact_uri,
    build_artifact_markdown_link,
)
from ..data_cleaning import CleanDatasetInput, DataCleaningService, cleaning_operation_metadata
from ..data_tokenization import DataTokenizationService, TokenizeDatasetInput
from ..data_transform import (
    DataQueryInput,
    DataQueryTransformService,
    DataTransformInput,
    DatasetSqlBinding,
)
from ..dataset_inspection import DatasetInspection, InspectDatasetInput, detect_source_format, load_dataframe
from ..dataset_service import DatasetService, RegisterDatasetInput
from ..ml.registry import get_model_catalog_entry, list_model_catalog, list_model_keys
from ..ml.types import EvaluationKind, ModelFamily, ModelTaskKind
from ..ml_service import (
    ApplyWithFilesInput,
    CreateColumnBindingInput,
    FitWithEvaluateInput,
    MLService,
    TuneWithEvaluateInput,
)
from ..storage.models import (
    ArtifactKind,
    DatasetSourceFormat,
    MLTaskArtifactKind,
    MLTaskRow,
    MLTaskStatus,
    MLTaskType,
    ProblemKind,
    TrainedModelRow,
)
from ..tabular import TabularSchema, resolve_tabular_schema, tabular_schema_payload
from ..llm import AgentToolSpec
from .tool_presentations import DEFAULT_TOOL_PRESENTATION, ToolPresentation, tool_presentation_for_name


_MODEL_ALIAS_SUFFIXES = {
    "classification",
    "classifier",
    "clustering",
    "regression",
    "regressor",
}
_MODEL_KEY_ALIAS_OVERRIDES = {
    "k_neighbors": "regression.knn",
    "kneighbors": "regression.knn",
    "k_neighbors_classifier": "classification.knn",
    "kneighborsclassifier": "classification.knn",
    "k_neighbors_regressor": "regression.knn",
    "kneighborsregressor": "regression.knn",
}
MODEL_APPLY_GRACE_SECONDS = 30.0
MODEL_TRAIN_GRACE_SECONDS = 60.0
MODEL_HYPER_TRAIN_GRACE_SECONDS = 60.0


@dataclass(frozen=True)
class ToolExecutionContext:
    thread_id: str
    turn_id: str
    tool_call_id: str
    dataset_ids: list[str]
    cancel_requested: Callable[[], bool] = lambda: False


class ToolExecutionResult(BaseModel):
    payload: dict[str, Any] = Field(default_factory=dict)
    content_blocks: list[dict[str, Any]] = Field(default_factory=list)
    provider_payload: dict[str, Any] | None = None

ToolHandler = Callable[[dict[str, Any], ToolExecutionContext], ToolExecutionResult]


@dataclass(frozen=True)
class AgentTool:
    spec: AgentToolSpec
    handler: ToolHandler
    presentation: ToolPresentation = DEFAULT_TOOL_PRESENTATION


class AgentToolRegistry:
    def __init__(
        self,
        *,
        paths: AppPaths,
        dataset_service: DatasetService,
        data_cleaning_service: DataCleaningService,
        data_transform_service: DataQueryTransformService,
        ml_service: MLService,
        artifact_service: ArtifactService,
        data_tokenization_service: DataTokenizationService | None = None,
        analysis_profile_service: AnalysisProfileService | None = None,
        analysis_graph_service: AnalysisGraphService | None = None,
        analysis_lambda_service: AnalysisLambdaService | None = None,
    ) -> None:
        self._paths = paths
        self._dataset_service = dataset_service
        self._data_cleaning_service = data_cleaning_service
        self._data_tokenization_service = data_tokenization_service or DataTokenizationService(paths)
        self._data_transform_service = data_transform_service
        self._analysis_profile_service = analysis_profile_service or AnalysisProfileService()
        self._analysis_graph_service = analysis_graph_service or AnalysisGraphService(paths)
        self._analysis_lambda_service = analysis_lambda_service or AnalysisLambdaService(paths)
        self._ml_service = ml_service
        self._artifact_service = artifact_service
        self._model_key_aliases = self._build_model_key_aliases()
        self._tools = {
            tool.spec.name: tool
            for tool in (
                self._build_data_peek_tool(),
                self._build_data_integrate_tool(),
                self._build_analysis_graph_tool(),
                # analysis.lambda is intentionally retained in code but not registered
                # in the Agent-facing tool set.
                # self._build_analysis_lambda_tool(),
                self._build_data_clean_tool(),
                self._build_data_clean_metadata_tool(),
                self._build_data_tokenize_tool(),
                self._build_data_query_tool(),
                self._build_data_transform_tool(),
                self._build_data_feature_select_tool(),
                self._build_model_metadata_tool(),
                self._build_model_train_tool(),
                self._build_model_hyper_train_tool(),
                self._build_model_apply_tool(),
                self._build_model_task_query_tool(),
            )
        }

    def list_specs(self) -> list[AgentToolSpec]:
        return [tool.spec for tool in self._tools.values()]

    def tool_presentation(self, tool_name: str) -> ToolPresentation:
        tool = self._tools.get(tool_name)
        if tool is None:
            return DEFAULT_TOOL_PRESENTATION
        return tool.presentation

    def execute(
        self,
        tool_name: str,
        arguments: dict[str, Any],
        context: ToolExecutionContext,
    ) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        tool = self._tools.get(tool_name)
        if tool is None:
            raise ValidationError(f"Tool '{tool_name}' is not registered.")
        result = tool.handler(arguments, context)
        self._raise_if_cancelled(context)
        return result

    def _build_data_peek_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="data.peek",
                provider_name="data_peek",
                description=(
                    "Inspect a registered dataset by dataset_id and by default "
                    "run bounded common descriptive analysis. Set analysis to false to inspect only."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {"type": "string"},
                        "analysis": {"type": "boolean", "default": True},
                        "target_columns": {"type": "array", "items": {"type": "string"}},
                        "top_n": {"type": "integer", "minimum": 1, "maximum": 20},
                        "correlation_column_limit": {"type": "integer", "minimum": 2, "maximum": 12},
                    },
                    "required": ["dataset_id"],
                    "additionalProperties": False,
                },
            ),
            handler=self._data_peek,
            presentation=tool_presentation_for_name("data.peek"),
        )

    def _build_data_integrate_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="data.integrate",
                provider_name="data_integrate",
                description="Combine two or more registered datasets into a generated dataset artifact.",
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_ids": {"type": "array", "items": {"type": "string"}, "minItems": 2},
                        "name": {"type": "string"},
                    },
                    "required": ["dataset_ids"],
                    "additionalProperties": False,
                },
            ),
            handler=self._data_integrate,
            presentation=tool_presentation_for_name("data.integrate"),
        )

    def _build_analysis_graph_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="analysis.graph",
                provider_name="analysis_graph",
                description=(
                    "Draw one bounded static SVG artifact for a registered dataset. Use exactly one of `spec` "
                    "or `wordcloud_spec`. `spec` is for ordinary Vega-Lite charts. "
                    "`wordcloud_spec` is the dedicated word-cloud path and expects an upstream chart-ready "
                    "frequency table from data.query or data.transform."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {
                            "type": "string",
                            "description": (
                                "Use one registered dataset. For word clouds, this dataset should already be a "
                                "chart-ready frequency table."
                            ),
                        },
                        "spec": {
                            "type": "object",
                            "description": (
                                "Vega-Lite chart specification. Xenix injects dataset values before rendering. "
                                "Use standard Vega-Lite mark, encoding, transform, layer, facet, concat, repeat, "
                                "config, and params fields. Do not include data or datasets, and do not use this "
                                "field for word clouds."
                            ),
                            "properties": {
                                "$schema": {"type": "string", "description": "Optional Vega-Lite schema URL."},
                                "width": {"type": "number", "description": "Chart width in pixels."},
                                "height": {"type": "number", "description": "Chart height in pixels."},
                                "title": {
                                    "description": "Optional chart title string or Vega-Lite title object.",
                                },
                                "mark": {
                                    "description": "Vega-Lite mark string or mark definition object.",
                                },
                                "encoding": {
                                    "type": "object",
                                    "description": "Vega-Lite channel encodings such as x, y, color, tooltip, and text.",
                                },
                                "transform": {
                                    "type": "array",
                                    "description": "Optional Vega-Lite transforms for lightweight chart shaping.",
                                },
                                "layer": {
                                    "type": "array",
                                    "description": "Optional Vega-Lite layers.",
                                },
                                "facet": {"type": "object", "description": "Optional Vega-Lite facet definition."},
                                "repeat": {"type": "object", "description": "Optional Vega-Lite repeat definition."},
                                "hconcat": {"type": "array", "description": "Optional horizontal concat views."},
                                "vconcat": {"type": "array", "description": "Optional vertical concat views."},
                                "concat": {"type": "array", "description": "Optional concat views."},
                                "spec": {"type": "object", "description": "Nested Vega-Lite view for facet/repeat."},
                                "config": {"type": "object", "description": "Optional Vega-Lite configuration."},
                                "params": {"type": "array", "description": "Optional Vega-Lite params."},
                            },
                            "additionalProperties": True,
                        },
                        "wordcloud_spec": {
                            "type": "object",
                            "description": (
                                "Dedicated word-cloud configuration. Use data.query or data.transform first to "
                                "produce chart-ready rows, usually exact columns `word` and `count`. For Chinese "
                                "text, segment upstream first; do not pass raw sentences or expect analysis.graph "
                                "to tokenize them."
                            ),
                            "properties": {
                                "title": {
                                    "type": "string",
                                    "description": "Optional visible word-cloud title.",
                                },
                                "word_field": {
                                    "type": "string",
                                    "description": "Word column name. Default is `word`.",
                                },
                                "count_field": {
                                    "type": "string",
                                    "description": "Positive count column name. Default is `count`.",
                                },
                                "top_n": {
                                    "type": "integer",
                                    "minimum": 20,
                                    "maximum": 80,
                                    "description": "Render only the top 20-80 terms for readability. Default is 80.",
                                },
                                "width": {
                                    "type": "integer",
                                    "minimum": 200,
                                    "maximum": 1600,
                                    "description": "Word-cloud width in pixels.",
                                },
                                "height": {
                                    "type": "integer",
                                    "minimum": 160,
                                    "maximum": 1200,
                                    "description": "Word-cloud height in pixels.",
                                },
                                "prefer_horizontal": {
                                    "type": "number",
                                    "minimum": 0.8,
                                    "maximum": 1.0,
                                    "description": "Keep at least 80% of terms horizontal. Default is 0.85.",
                                },
                                "font_size_range": {
                                    "type": "array",
                                    "minItems": 2,
                                    "maxItems": 2,
                                    "items": {"type": "number"},
                                    "description": (
                                        "Optional [min, max] font-size range. Default is [12, 56], or [10, 42] "
                                        "for denser clouds."
                                    ),
                                },
                                "color_mode": {
                                    "type": "string",
                                    "enum": ["rank_tier", "field"],
                                    "description": (
                                        "Use `rank_tier` for restrained 2-3 color ranking. Use `field` only when "
                                        "an upstream low-cardinality field already encodes category, sentiment, or source."
                                    ),
                                },
                                "color_field": {
                                    "type": "string",
                                    "description": "Required only when color_mode is `field`.",
                                },
                                "palette": {
                                    "type": "array",
                                    "items": {"type": "string"},
                                    "description": "Optional restrained color palette for rank tiers or semantic groups.",
                                },
                            },
                            "additionalProperties": False,
                        },
                    },
                    "required": ["dataset_id"],
                    "oneOf": [
                        {"required": ["spec"]},
                        {"required": ["wordcloud_spec"]},
                    ],
                    "additionalProperties": False,
                },
            ),
            handler=self._analysis_graph,
            presentation=tool_presentation_for_name("analysis.graph"),
        )

    def _build_analysis_lambda_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="analysis.lambda",
                provider_name="analysis_lambda",
                description=(
                    "Run a one-off Python analysis function over registered datasets. "
                    "The code must define analyze(ctx, inputs, params) and return any JSON-serializable dict. "
                    "inputs is a mapping from dataset alias to pandas DataFrame; inputs[alias].read() also returns "
                    "that DataFrame. Supported imports: pandas/pd, numpy/np, matplotlib/plt, scipy, statsmodels, "
                    "sklearn, xgboost, lightgbm, math, statistics, datetime, json, io, collections, itertools, "
                    "functools, and typing. Do not import seaborn or arbitrary packages. Use ctx.artifact.create(...) "
                    "for generated artifacts: ctx.artifact.create(name, content), ctx.artifact.create(content, name=...), "
                    "or ctx.artifact.create(name=..., content=...); content may be a pandas DataFrame, SVG/text string, "
                    "bytes/io.BytesIO, or matplotlib Figure. value=... is accepted as an alias for content=...."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "code": {"type": "string"},
                        "datasets": {
                            "type": "object",
                            "additionalProperties": {"type": "string"},
                            "description": "Mapping from dataset alias to registered dataset_id.",
                        },
                        "params": {"type": "object"},
                        "manifest": {"type": "object"},
                    },
                    "required": ["code", "datasets"],
                    "additionalProperties": False,
                },
            ),
            handler=self._analysis_lambda,
            presentation=tool_presentation_for_name("analysis.lambda"),
        )

    def _build_data_clean_tool(self) -> AgentTool:
        cleaning_operation_schema = {
            "type": "object",
            "properties": {
                "operation": {"type": "string"},
                "params": {"type": "object"},
            },
            "required": ["operation"],
            "additionalProperties": False,
        }
        return AgentTool(
            spec=AgentToolSpec(
                name="data.clean",
                provider_name="data_clean",
                description=(
                    "Create a new derived dataset by applying atomic predefined cleaning operations "
                    "to one registered dataset. Call data.clean.metadata for operation parameter schemas."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {"type": "string"},
                        "name": {"type": "string"},
                        "operations": {"type": "array", "items": cleaning_operation_schema},
                    },
                    "required": ["dataset_id"],
                    "additionalProperties": False,
                },
            ),
            handler=self._data_clean,
            presentation=tool_presentation_for_name("data.clean"),
        )

    def _build_data_clean_metadata_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="data.clean.metadata",
                provider_name="data_clean_metadata",
                description="Return data.clean operation groups, operation names, and parameter schemas.",
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "groups": {"type": "array", "items": {"type": "string"}},
                    },
                    "additionalProperties": False,
                },
            ),
            handler=self._data_clean_metadata,
            presentation=tool_presentation_for_name("data.clean.metadata"),
        )

    def _build_data_tokenize_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="data.tokenize",
                provider_name="data_tokenize",
                description=(
                    "Create a new derived dataset by tokenizing one Chinese text column with the stable "
                    "Xenix profile zh_business_v1. Use output=token_text to keep source rows and append "
                    "token_text for downstream text models, or output=token_rows to explode one token per row."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {"type": "string"},
                        "name": {"type": "string"},
                        "text_column": {
                            "type": "string",
                            "description": "Chinese text column to segment.",
                        },
                        "id_columns": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Optional identifier columns preserved in token_rows output.",
                        },
                        "output": {
                            "type": "string",
                            "enum": ["token_text", "token_rows"],
                            "description": "Choose token_text for model-ready rows or token_rows for one-token-per-row analysis.",
                        },
                        "tokenizer_profile": {
                            "type": "string",
                            "enum": ["zh_business_v1"],
                            "description": "Stable Chinese-first tokenization profile owned by Xenix.",
                        },
                    },
                    "required": ["dataset_id", "text_column"],
                    "additionalProperties": False,
                },
            ),
            handler=self._data_tokenize,
            presentation=tool_presentation_for_name("data.tokenize"),
        )

    def _build_data_query_tool(self) -> AgentTool:
        binding_schema = {
            "type": "object",
            "properties": {
                "alias": {
                    "type": "string",
                    "description": "SQL table alias for this registered dataset, such as orders or customers.",
                },
                "dataset_id": {
                    "type": "string",
                    "description": "Registered dataset id bound to this SQL alias.",
                },
            },
            "required": ["alias", "dataset_id"],
        }
        return AgentTool(
            spec=AgentToolSpec(
                name="data.query",
                provider_name="data_query",
                description=(
                    "Run a read-only SELECT/CTE query over registered datasets. "
                    "Use dataset_id for one input aliased as input, or bindings for multiple inputs. "
                    "Returns bounded rows and does not create a dataset artifact."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {
                            "type": "string",
                            "description": "Use for one input dataset, which will be available in SQL as input.",
                        },
                        "bindings": {
                            "type": "array",
                            "items": binding_schema,
                            "description": "Use for multiple input datasets with explicit SQL aliases.",
                        },
                        "sql": {
                            "type": "string",
                            "description": "Read-only SELECT or CTE query to execute.",
                        },
                        "limit": {
                            "type": "integer",
                            "minimum": 1,
                            "maximum": 200,
                            "description": "Maximum number of rows to return in the bounded result.",
                        },
                    },
                    "required": ["sql"],
                },
            ),
            handler=self._data_query,
            presentation=tool_presentation_for_name("data.query"),
        )

    def _build_data_transform_tool(self) -> AgentTool:
        binding_schema = {
            "type": "object",
            "properties": {
                "alias": {
                    "type": "string",
                    "description": "SQL table alias for this registered dataset, such as orders or customers.",
                },
                "dataset_id": {
                    "type": "string",
                    "description": "Registered dataset id bound to this SQL alias.",
                },
            },
            "required": ["alias", "dataset_id"],
        }
        return AgentTool(
            spec=AgentToolSpec(
                name="data.transform",
                provider_name="data_transform",
                description=(
                    "Create a new derived dataset artifact from a SELECT/CTE query over registered datasets. "
                    "Use dataset_id for one input aliased as input, or bindings for multiple inputs."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {
                            "type": "string",
                            "description": "Use for one input dataset, which will be available in SQL as input.",
                        },
                        "bindings": {
                            "type": "array",
                            "items": binding_schema,
                            "description": "Use for multiple input datasets with explicit SQL aliases.",
                        },
                        "sql": {
                            "type": "string",
                            "description": "SELECT or CTE query whose result becomes the generated dataset.",
                        },
                        "name": {
                            "type": "string",
                            "description": "Optional name for the generated transformed dataset.",
                        },
                    },
                    "required": ["sql"],
                },
            ),
            handler=self._data_transform,
            presentation=tool_presentation_for_name("data.transform"),
        )

    def _build_data_feature_select_tool(self) -> AgentTool:
        role_binding_schema = {
            "type": "object",
            "properties": {
                "role": {
                    "type": "string",
                    "description": "Semantic role such as feature, target, or partial_target.",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Dataset columns assigned to this semantic role.",
                },
            },
            "required": ["role", "columns"],
        }
        return AgentTool(
            spec=AgentToolSpec(
                name="data.feature.select",
                provider_name="data_feature_select",
                description="Bind registered dataset columns to semantic roles required by a model/analyzer.",
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "dataset_id": {
                            "type": "string",
                            "description": "Registered dataset id whose columns will be role-bound.",
                        },
                        "model_key": {
                            "type": "string",
                            "description": "Optional chosen model key to pre-align role validation and later training.",
                        },
                        "role_bindings": {
                            "type": "array",
                            "items": role_binding_schema,
                            "description": "Semantic role bindings to persist for later model training or apply.",
                        },
                    },
                    "required": ["dataset_id", "role_bindings"],
                },
            ),
            handler=self._data_feature_select,
            presentation=tool_presentation_for_name("data.feature.select"),
        )

    def _build_model_metadata_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="model.metadata",
                provider_name="model_metadata",
                description=(
                    "Browse a lightweight model directory by model_family, or inspect one chosen model's role "
                    "and parameter schema with model_key."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "model_key": {
                            "type": "string",
                            "description": (
                                "Inspect one chosen model. Accepts a canonical model key or a simple alias. "
                                "Returns role schemas and param_schema by default."
                            ),
                        },
                        "model_family": {
                            "type": "string",
                            "enum": [family.value for family in ModelFamily],
                            "description": (
                                "Browse lightweight candidate models in one family such as supervised, "
                                "clustering, anomaly_detection, association_rules, recommendation, "
                                "or text_analysis."
                            ),
                        },
                        "include_param_grid_schema": {
                            "type": "boolean",
                            "description": (
                                "Only use with model_key. When true, also return param_grid_schema for "
                                "hyperparameter tuning."
                            ),
                        },
                    },
                },
            ),
            handler=self._model_metadata,
            presentation=tool_presentation_for_name("model.metadata"),
        )

    def _build_model_train_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="model.train",
                provider_name="model_train",
                description=(
                    "Train and evaluate one or more models for a persisted dataset column role binding. "
                    "Use model.metadata with model_family to browse candidates, then inspect one model_key for "
                    "parameter detail."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "binding_id": {
                            "type": "string",
                            "description": "Column role-binding id returned by data.feature.select.",
                        },
                        "models": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "One or more chosen model keys or aliases to train.",
                        },
                        "params_by_model": {
                            "type": "object",
                            "description": "Optional per-model parameter objects keyed by model key.",
                        },
                        "run_name": {
                            "type": "string",
                            "description": "Optional human-readable run name.",
                        },
                    },
                    "required": ["binding_id", "models"],
                },
            ),
            handler=self._model_train,
            presentation=tool_presentation_for_name("model.train"),
        )

    def _build_model_hyper_train_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="model.hyper_train",
                provider_name="model_hyper_train",
                description=(
                    "Run hyperparameter training for one or more models. "
                    "Use model.metadata with model_family to browse candidates, then inspect one model_key with "
                    "include_param_grid_schema=true."
                ),
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "binding_id": {
                            "type": "string",
                            "description": "Column role-binding id returned by data.feature.select.",
                        },
                        "param_grids_by_model": {
                            "type": "object",
                            "description": "Per-model tuning grids keyed by model key.",
                        },
                        "run_name": {
                            "type": "string",
                            "description": "Optional human-readable run name.",
                        },
                    },
                    "required": ["binding_id", "param_grids_by_model"],
                },
            ),
            handler=self._model_hyper_train,
            presentation=tool_presentation_for_name("model.hyper_train"),
        )

    def _build_model_apply_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="model.apply",
                provider_name="model_apply",
                description="Apply a trained model to registered dataset or artifact inputs, or inline rows.",
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "trained_model_id": {"type": "string"},
                        "input_sources": {"type": "array", "items": {"type": "string"}},
                        "input_rows": {
                            "type": "object",
                            "properties": {
                                "header_index_map": {
                                    "type": "object",
                                    "additionalProperties": {"type": "integer", "minimum": 0},
                                },
                                "data": {
                                    "type": "array",
                                    "items": {
                                        "type": "array",
                                        "items": {
                                            "type": ["string", "number", "boolean", "null"],
                                        },
                                    },
                                },
                            },
                            "required": ["header_index_map", "data"],
                            "additionalProperties": False,
                        },
                    },
                    "required": ["trained_model_id"],
                    "additionalProperties": False,
                },
            ),
            handler=self._model_apply,
            presentation=tool_presentation_for_name("model.apply"),
        )

    def _build_model_task_query_tool(self) -> AgentTool:
        return AgentTool(
            spec=AgentToolSpec(
                name="model.task.query",
                provider_name="model_task_query",
                description="Query ML task status, metadata, artifacts, errors, and logs by explicit task ids.",
                parameters_schema={
                    "type": "object",
                    "properties": {
                        "task_ids": {
                            "type": "array",
                            "items": {"type": "string"},
                            "minItems": 1,
                            "description": "One or more explicit ML task ids to inspect.",
                        },
                        "include_logs": {
                            "type": "boolean",
                            "description": "Set true to include bounded task logs in the response.",
                        },
                        "max_log_entries": {
                            "type": "integer",
                            "minimum": 0,
                            "maximum": 1000,
                            "description": "Maximum number of log entries to return per task when include_logs is true.",
                        },
                    },
                    "required": ["task_ids"],
                },
            ),
            handler=self._model_task_query,
            presentation=tool_presentation_for_name("model.task.query"),
        )

    def _data_peek(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        supported_keys = {
            "dataset_id",
            "analysis",
            "target_columns",
            "top_n",
            "correlation_column_limit",
        }
        unsupported_keys = sorted(set(arguments) - supported_keys)
        if unsupported_keys:
            raise ValidationError("data.peek does not accept: " + ", ".join(unsupported_keys))
        analysis_enabled = self._optional_boolean(arguments, "analysis", default=True)
        dataset_id = self._require_string(arguments, "dataset_id")
        dataset = self._dataset_service.get_dataset(dataset_id)
        inspection = self._dataset_service.inspect_source_file(
            InspectDatasetInput(source_path=dataset.source_path)
        )
        payload: dict[str, Any] = {
            "dataset_id": dataset.id,
            "inspection": self._inspection_payload(inspection),
        }
        structure = self._structure_payload(Path(dataset.source_path).expanduser(), inspection)
        payload["structure"] = structure
        markdown_text = (
            f"Dataset `{dataset.name}` is ready.\n\n"
            f"Rows: {inspection.row_count}; columns: {', '.join(inspection.preview_columns)}"
        )
        if analysis_enabled:
            profile_result = self._analysis_profile_service.profile_dataset(
                ProfileDatasetInput(
                    source_path=dataset.source_path,
                    dataset_name=dataset.name,
                    target_columns=self._optional_string_list(arguments, "target_columns"),
                    top_n=self._optional_integer(arguments, "top_n", default=10),
                    correlation_column_limit=self._optional_integer(
                        arguments,
                        "correlation_column_limit",
                        default=8,
                    ),
                )
            )
            payload["analysis"] = {
                "enabled": True,
                "profile": profile_result.profile,
                "markdown": profile_result.markdown,
            }
            markdown_text = f"{markdown_text}\n\n{profile_result.markdown}"
        else:
            payload["analysis"] = {"enabled": False}
        provider_payload = self._data_peek_provider_payload(payload)
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": markdown_text}],
            provider_payload=provider_payload,
        )

    def _data_integrate(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        raw_dataset_ids = arguments.get("dataset_ids")
        if not isinstance(raw_dataset_ids, list) or len(raw_dataset_ids) < 2:
            raise ValidationError("data.integrate requires at least two dataset ids.")
        datasets = [self._dataset_service.get_dataset(str(dataset_id)) for dataset_id in raw_dataset_ids]
        frames = [self._load_frame(Path(dataset.source_path).expanduser().resolve()) for dataset in datasets]
        output_dir = self._paths.artifacts / "datasets" / "integrated"
        output_dir.mkdir(parents=True, exist_ok=True)
        name = str(arguments.get("name") or "Integrated dataset").strip() or "Integrated dataset"
        output_path = output_dir / f"{self._slug(name)}-{int(time.time())}.csv"
        pd.concat(frames, ignore_index=True).to_csv(output_path, index=False)
        input_dataset_ids = [dataset.id for dataset in datasets]
        result = self._register_generated_dataset_result(
            context,
            output_path=output_path,
            name=name,
            summary="Integrated dataset created.",
            metadata_payload={"input_dataset_ids": input_dataset_ids},
        )
        result.payload["input_dataset_ids"] = input_dataset_ids
        return result

    def _analysis_graph(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        unsupported_keys = sorted(set(arguments) - {"dataset_id", "spec", "wordcloud_spec"})
        if unsupported_keys:
            raise ValidationError("analysis.graph does not accept: " + ", ".join(unsupported_keys))
        dataset_id = self._require_string(arguments, "dataset_id")
        dataset = self._dataset_service.get_dataset(dataset_id)
        has_spec = "spec" in arguments
        has_wordcloud_spec = "wordcloud_spec" in arguments
        if has_spec == has_wordcloud_spec:
            raise ValidationError("analysis.graph requires exactly one of spec or wordcloud_spec.")
        raw_spec = arguments.get("spec")
        raw_wordcloud_spec = arguments.get("wordcloud_spec")
        if has_spec and not isinstance(raw_spec, dict):
            raise ValidationError("analysis.graph spec must be a Vega-Lite object.")
        if has_wordcloud_spec and not isinstance(raw_wordcloud_spec, dict):
            raise ValidationError("analysis.graph wordcloud_spec must be an object.")
        graph_result = self._analysis_graph_service.graph_dataset(
            GraphDatasetInput(
                source_path=dataset.source_path,
                dataset_name=dataset.name,
                spec=raw_spec,
                wordcloud_spec=raw_wordcloud_spec,
            )
        )
        graph_metadata = graph_result.graph_metadata
        default_title = f"{dataset.name} graph"
        title = str(graph_metadata.get("title") or default_title).strip() or default_title
        spec_format = str(graph_metadata.get("spec_format") or "graph")
        artifact = self._artifact_service.register_artifact(
            RegisterArtifactInput(
                thread_id=context.thread_id,
                turn_id=context.turn_id,
                tool_call_id=context.tool_call_id,
                kind=ArtifactKind.IMAGE,
                title=title,
                absolute_path=graph_result.output_path,
                mime_type="image/svg+xml",
                summary=f"Graph generated by analysis.graph ({spec_format}).",
                preview_payload=graph_metadata,
                metadata_payload={
                    "dataset_id": dataset.id,
                    "analysis_graph": graph_metadata,
                },
            )
        )
        payload = {
            "dataset_id": dataset.id,
            "artifact_id": artifact.id,
            "graph": graph_metadata,
        }
        return ToolExecutionResult(payload=payload)

    def _analysis_lambda(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        unsupported_keys = sorted(set(arguments) - {"code", "datasets", "params", "manifest"})
        if unsupported_keys:
            raise ValidationError("analysis.lambda does not accept: " + ", ".join(unsupported_keys))
        raw_datasets = arguments.get("datasets")
        if not isinstance(raw_datasets, dict) or not raw_datasets:
            raise ValidationError("analysis.lambda datasets must be a non-empty object.")
        params = arguments.get("params") or {}
        if not isinstance(params, dict):
            raise ValidationError("analysis.lambda params must be an object.")
        manifest = arguments.get("manifest") or {}
        if not isinstance(manifest, dict):
            raise ValidationError("analysis.lambda manifest must be an object.")

        datasets: list[AnalysisLambdaDataset] = []
        for raw_alias, raw_dataset_id in raw_datasets.items():
            alias = str(raw_alias or "").strip()
            dataset_id = str(raw_dataset_id or "").strip()
            if not alias or not dataset_id:
                raise ValidationError("analysis.lambda datasets must map non-empty aliases to dataset ids.")
            dataset = self._dataset_service.get_dataset(dataset_id)
            datasets.append(
                AnalysisLambdaDataset(
                    alias=alias,
                    dataset_id=dataset.id,
                    dataset_name=dataset.name,
                    source_path=dataset.source_path,
                )
            )

        lambda_result = self._analysis_lambda_service.run_lambda(
            AnalysisLambdaInput(
                code=self._require_string(arguments, "code"),
                datasets=datasets,
                params=params,
                manifest=manifest,
            ),
            cancel_requested=context.cancel_requested,
        )
        artifact_map: dict[str, str] = {}
        artifact_payloads: list[dict[str, Any]] = []
        for descriptor in lambda_result.artifacts:
            kind = self._lambda_artifact_kind(descriptor.kind)
            metadata_payload = {
                "analysis_lambda": {
                    "placeholder_id": descriptor.placeholder_id,
                    "kind": descriptor.kind,
                    "metadata": descriptor.metadata_payload,
                }
            }
            artifact = self._artifact_service.register_artifact(
                RegisterArtifactInput(
                    thread_id=context.thread_id,
                    turn_id=context.turn_id,
                    tool_call_id=context.tool_call_id,
                    kind=kind,
                    title=descriptor.title,
                    absolute_path=descriptor.absolute_path,
                    mime_type=descriptor.mime_type,
                    summary=descriptor.summary,
                    metadata_payload=metadata_payload,
                )
            )
            uri = build_artifact_uri(artifact.id)
            artifact_map[descriptor.placeholder_id] = artifact.id
            artifact_payloads.append(
                {
                    "artifact_id": artifact.id,
                    "uri": uri,
                    "title": artifact.title,
                    "kind": artifact.kind.value,
                    "mime_type": artifact.mime_type,
                    "placeholder_id": descriptor.placeholder_id,
                }
            )

        output = self._rewrite_lambda_artifact_uris(lambda_result.output, artifact_map)
        payload = {
            "result": {
                "output": output,
            },
            "artifacts": artifact_payloads,
            "dataset_ids": [dataset.dataset_id for dataset in datasets],
        }
        content_blocks = []
        markdown = output.get("markdown")
        if isinstance(markdown, str) and markdown.strip():
            content_blocks.append({"type": "markdown", "text": markdown})
        return ToolExecutionResult(payload=payload, content_blocks=content_blocks)

    def _data_clean(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        unsupported_keys = sorted(set(arguments) - {"dataset_id", "name", "operations"})
        if unsupported_keys:
            raise ValidationError("data.clean does not accept: " + ", ".join(unsupported_keys))
        dataset_id = self._require_string(arguments, "dataset_id")
        dataset = self._dataset_service.get_dataset(dataset_id)
        name = str(arguments.get("name") or f"{dataset.name} cleaned").strip() or f"{dataset.name} cleaned"
        operations = arguments.get("operations") or []
        if not isinstance(operations, list):
            raise ValidationError("operations must be a list.")
        if not operations:
            report = {
                "row_count_before": None,
                "row_count_after": None,
                "rows_removed": 0,
                "operations": [],
                "validation_rules": [],
                "warnings": [],
                "no_op": True,
            }
            return ToolExecutionResult(
                payload={
                    "dataset_id": dataset.id,
                    "source_dataset_id": dataset.id,
                    "cleaning_report": report,
                    "message": "No cleaning operations were requested. Nothing happened.",
                },
                content_blocks=[
                    {
                        "type": "markdown",
                        "text": "No cleaning operations were requested. Nothing happened; the source dataset was left unchanged.",
                    }
                ],
            )
        try:
            clean_input = CleanDatasetInput(
                source_path=dataset.source_path,
                name=name,
                operations=operations,
            )
        except PydanticValidationError as exc:
            raise ValidationError("operations must contain objects with operation and optional params.") from exc
        clean_result = self._data_cleaning_service.clean_dataset(clean_input)
        row_count_before = int(clean_result.report.get("row_count_before", 0))
        row_count_after = int(clean_result.report.get("row_count_after", 0))
        result = self._register_generated_dataset_result(
            context,
            output_path=Path(clean_result.output_path),
            name=name,
            summary=f"Cleaned dataset created. Rows: {row_count_before} -> {row_count_after}.",
            derived_from_dataset_id=dataset.id,
            metadata_payload={"cleaning_report": clean_result.report},
        )
        result.payload["row_count_before"] = row_count_before
        result.payload["row_count_after"] = row_count_after
        result.payload["cleaning_report"] = clean_result.report
        return result

    def _data_clean_metadata(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        unsupported_keys = sorted(set(arguments) - {"groups"})
        if unsupported_keys:
            raise ValidationError("data.clean.metadata does not accept: " + ", ".join(unsupported_keys))
        groups = self._optional_string_list(arguments, "groups")
        payload = cleaning_operation_metadata(groups)
        group_names = ", ".join(payload["group_names"])
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[
                {
                    "type": "markdown",
                    "text": f"Available data.clean operation groups: {group_names}.",
                }
            ],
        )

    def _data_tokenize(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        unsupported_keys = sorted(set(arguments) - {"dataset_id", "name", "text_column", "id_columns", "output", "tokenizer_profile"})
        if unsupported_keys:
            raise ValidationError("data.tokenize does not accept: " + ", ".join(unsupported_keys))
        dataset_id = self._require_string(arguments, "dataset_id")
        dataset = self._dataset_service.get_dataset(dataset_id)
        default_name = f"{dataset.name} tokenized"
        name = str(arguments.get("name") or default_name).strip() or default_name
        id_columns = arguments.get("id_columns") or []
        if not isinstance(id_columns, list) or not all(isinstance(item, str) for item in id_columns):
            raise ValidationError("data.tokenize id_columns must be a list of strings.")
        tokenize_result = self._data_tokenization_service.tokenize_dataset(
            TokenizeDatasetInput(
                source_path=dataset.source_path,
                name=name,
                text_column=self._require_string(arguments, "text_column"),
                id_columns=[str(item) for item in id_columns],
                output=str(arguments.get("output") or "token_text"),
                tokenizer_profile=str(arguments.get("tokenizer_profile") or "zh_business_v1"),
            )
        )
        row_count = int(tokenize_result.report.get("output_row_count", 0))
        result = self._register_generated_dataset_result(
            context,
            output_path=Path(tokenize_result.output_path),
            name=name,
            summary=f"Tokenized dataset created. Rows: {row_count}.",
            derived_from_dataset_id=dataset.id,
            metadata_payload={"tokenization_report": tokenize_result.report},
        )
        result.payload["row_count"] = row_count
        result.payload["tokenization_report"] = tokenize_result.report
        return result

    def _data_query(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        bindings = self._resolve_sql_bindings(arguments, tool_name="data.query")
        query_result = self._data_transform_service.query(
            DataQueryInput(
                bindings=bindings,
                sql=self._require_string(arguments, "sql"),
                limit=self._optional_integer(arguments, "limit", default=50),
            )
        )
        payload = query_result.model_dump(mode="json")
        payload["input_dataset_ids"] = [binding.dataset_id for binding in bindings]
        payload["bindings"] = [
            {"alias": binding.alias, "dataset_id": binding.dataset_id}
            for binding in bindings
        ]
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": self._query_result_markdown(payload)}],
        )

    def _data_transform(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        bindings = self._resolve_sql_bindings(arguments, tool_name="data.transform")
        input_dataset_ids = [binding.dataset_id for binding in bindings]
        default_name = "Transformed dataset"
        if len(bindings) == 1:
            default_name = f"{self._dataset_service.get_dataset(bindings[0].dataset_id).name} transformed"
        name = str(arguments.get("name") or default_name).strip() or default_name
        transform_result = self._data_transform_service.transform(
            DataTransformInput(
                bindings=bindings,
                sql=self._require_string(arguments, "sql"),
                name=name,
            )
        )
        derived_from_dataset_id = input_dataset_ids[0] if len(set(input_dataset_ids)) == 1 else None
        result = self._register_generated_dataset_result(
            context,
            output_path=Path(transform_result.output_path),
            name=name,
            summary=f"Transformed dataset created. Rows: {transform_result.row_count}.",
            derived_from_dataset_id=derived_from_dataset_id,
            metadata_payload={
                "transform_report": transform_result.transform_report,
                "input_dataset_ids": input_dataset_ids,
            },
        )
        result.payload["row_count"] = transform_result.row_count
        result.payload["columns"] = transform_result.columns
        result.payload["transform_report"] = transform_result.transform_report
        result.payload["input_dataset_ids"] = input_dataset_ids
        return result

    def _data_feature_select(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        dataset_id = self._require_string(arguments, "dataset_id")
        role_bindings = arguments.get("role_bindings")
        if not isinstance(role_bindings, list):
            raise ValidationError("data.feature.select requires role_bindings.")
        if not all(isinstance(item, dict) for item in role_bindings):
            raise ValidationError("data.feature.select role_bindings must contain objects.")
        binding = self._ml_service.create_column_binding(
            CreateColumnBindingInput(
                dataset_id=dataset_id,
                model_key=str(arguments.get("model_key") or "").strip() or None,
                role_bindings=[dict(item) for item in role_bindings],
            )
        )
        return ToolExecutionResult(
            payload={
                "binding_id": binding.id,
                "dataset_id": binding.dataset_id,
                "role_bindings": list(binding.role_bindings),
                "model_key": binding.model_key,
                "model_family": binding.model_family,
                "model_task_kind": binding.model_task_kind,
            },
            content_blocks=[
                {
                    "type": "markdown",
                    "text": (
                        f"Binding id: `{binding.id}`\n\n"
                        f"Bound roles: {', '.join(str(item.get('role')) for item in binding.role_bindings)}"
                    ),
                }
            ],
        )

    def _model_metadata(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        raw_model_key = str(arguments.get("model_key") or "").strip()
        raw_model_keys = self._optional_string_list(arguments, "model_keys")
        if raw_model_key and raw_model_keys:
            raise ValidationError("model.metadata accepts either model_key or model_keys, not both.")

        detail_model_key: str | None = None
        if raw_model_key:
            detail_model_key = self._normalize_model_keys([raw_model_key], field_name="model_key")[0]
        elif raw_model_keys:
            model_keys = self._normalize_model_keys(raw_model_keys, field_name="model_keys")
            if len(model_keys) != 1:
                raise ValidationError(
                    "model.metadata accepts only one model_key for parameter detail. "
                    "Use narrowing filters such as model_family for directory discovery."
                )
            detail_model_key = model_keys[0]

        model_family = str(arguments.get("model_family") or "").strip()
        if model_family:
            try:
                selected_model_family = ModelFamily(model_family)
            except ValueError as exc:
                raise ValidationError(f"Unknown model_family '{model_family}'.") from exc

        detail_query = detail_model_key is not None
        has_directory_filter = bool(model_family)
        if not detail_query and not has_directory_filter:
            raise ValidationError(
                "model.metadata requires model_key or model_family."
            )

        if detail_query:
            catalog_entries = [get_model_catalog_entry(detail_model_key)]
        else:
            catalog_entries = list_model_catalog()

        if model_family:
            catalog_entries = [
                entry for entry in catalog_entries if entry.model_family == selected_model_family
            ]

        catalog_entries = sorted(
            catalog_entries,
            key=lambda entry: (
                entry.model_family.value,
                entry.model_task_kind.value,
                entry.evaluation_kind.value,
                entry.problem_kind.value if entry.problem_kind is not None else "",
                entry.recommendation_tier,
                entry.model_key,
            ),
        )
        include_param_grid_schema = self._optional_boolean(
            arguments,
            "include_param_grid_schema",
            default=False,
        )
        legacy_include_param_schema = self._optional_boolean(
            arguments,
            "include_param_schema",
            default=False,
        )
        include_param_schema = detail_query or include_param_grid_schema or legacy_include_param_schema
        if not detail_query:
            include_param_schema = False
            include_param_grid_schema = False
        models = [
            self._model_catalog_payload(
                entry,
                detail_query=detail_query,
                include_param_schema=include_param_schema,
                include_param_grid_schema=include_param_grid_schema,
            )
            for entry in catalog_entries
        ]
        payload = {
            "model_keys": [model["model_key"] for model in models],
            "models": models,
        }
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": self._model_metadata_markdown(models)}],
        )

    def _model_train(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        binding_id = self._require_string(arguments, "binding_id")
        models = self._normalize_model_keys(
            self._require_string_list(arguments, "models"),
            field_name="models",
        )
        params_by_model = self._normalize_model_mapping(
            arguments.get("params_by_model"),
            field_name="params_by_model",
        )
        binding = self._ml_service.get_column_binding(binding_id)
        dataset_id = binding.dataset_id
        created_task_ids: list[str] = []
        for model_key in models:
            self._raise_if_cancelled(context)
            created = self._ml_service.fit_with_evaluate(
                FitWithEvaluateInput(
                    binding_id=binding_id,
                    run_name=str(arguments.get("run_name") or ""),
                    model_key=model_key,
                    params=dict(params_by_model.get(model_key) or {}),
                )
            )
            created_task_ids.append(created.id)
        training_result = self._wait_for_training_models_or_none(
            created_task_ids,
            context=context,
            timeout_seconds=MODEL_TRAIN_GRACE_SECONDS,
        )
        if training_result is None:
            return self._training_task_receipt(
                tool_name="model.train",
                dataset_id=dataset_id,
                root_task_ids=created_task_ids,
                operation="fit",
            )
        tasks, trained_models = training_result
        payload = {
            "async_state": "completed",
            "dataset_id": dataset_id,
            "task_ids": [task.id for task in tasks],
            "ml_tasks": [self._ml_task_payload(task) for task in tasks],
            "trained_models": [self._trained_model_payload(model) for model in trained_models],
        }
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": self._training_summary_markdown(payload)}],
        )

    def _model_hyper_train(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        binding_id = self._require_string(arguments, "binding_id")
        grids = arguments.get("param_grids_by_model")
        if not isinstance(grids, dict) or not grids:
            raise ValidationError("model.hyper_train requires param_grids_by_model.")
        normalized_grids = self._normalize_model_mapping(
            grids,
            field_name="param_grids_by_model",
            require_hyperparameter_tuning=True,
        )
        binding = self._ml_service.get_column_binding(binding_id)
        dataset_id = binding.dataset_id
        created_task_ids: list[str] = []
        for model_key, grid in normalized_grids.items():
            self._raise_if_cancelled(context)
            created = self._ml_service.tune_with_evaluate(
                TuneWithEvaluateInput(
                    binding_id=binding_id,
                    run_name=str(arguments.get("run_name") or ""),
                    model_key=model_key,
                    param_grid=dict(grid),
                )
            )
            created_task_ids.append(created.id)
        training_result = self._wait_for_training_models_or_none(
            created_task_ids,
            context=context,
            timeout_seconds=MODEL_HYPER_TRAIN_GRACE_SECONDS,
        )
        if training_result is None:
            return self._training_task_receipt(
                tool_name="model.hyper_train",
                dataset_id=dataset_id,
                root_task_ids=created_task_ids,
                operation="hyperparameter_tuning",
            )
        tasks, trained_models = training_result
        payload = {
            "async_state": "completed",
            "dataset_id": dataset_id,
            "task_ids": [task.id for task in tasks],
            "ml_tasks": [self._ml_task_payload(task) for task in tasks],
            "trained_models": [self._trained_model_payload(model) for model in trained_models],
        }
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": self._training_summary_markdown(payload)}],
        )

    def _model_apply(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        trained_model_id = self._require_string(arguments, "trained_model_id")
        resolved_input_files = self._resolve_apply_input_sources(self._optional_string_list(arguments, "input_sources"))
        input_rows = arguments.get("input_rows")
        if input_rows is not None and not isinstance(input_rows, dict):
            raise ValidationError("input_rows must be an object.")
        if not resolved_input_files and input_rows is None:
            raise ValidationError("model.apply requires input_sources or input_rows.")
        try:
            apply_input = ApplyWithFilesInput(
                trained_model_id=trained_model_id,
                input_files=resolved_input_files,
                input_rows=input_rows,
            )
        except PydanticValidationError as exc:
            raise ValidationError("input_rows must contain header_index_map and data.") from exc
        task = self._ml_service.apply(apply_input)
        completed_task = self._wait_for_task_or_none(
            task.id,
            context=context,
            timeout_seconds=MODEL_APPLY_GRACE_SECONDS,
        )
        if completed_task is None:
            return self._single_task_receipt(
                tool_name="model.apply",
                task_id=task.id,
                operation="apply",
            )
        task = completed_task
        details = self._ml_service.get_task_details(task.id)
        output_artifact = next(
            artifact
            for artifact in details.artifacts
            if artifact.artifact_kind is MLTaskArtifactKind.APPLY_RESULT
        )
        generic_artifact = self._artifact_service.register_artifact(
            RegisterArtifactInput(
                thread_id=context.thread_id,
                turn_id=context.turn_id,
                tool_call_id=context.tool_call_id,
                kind=ArtifactKind.FILE,
                title="Apply results",
                absolute_path=output_artifact.absolute_path,
                mime_type="text/csv",
                metadata_payload={"ml_task_id": task.id, "dataset_id": task.dataset_id},
            )
        )
        link = build_artifact_markdown_link(generic_artifact)
        return ToolExecutionResult(
            payload={
                "async_state": "completed",
                "ml_task_id": task.id,
                "task_ids": [task.id],
                "ml_tasks": [self._ml_task_payload(task)],
                "dataset_id": task.dataset_id,
                "result_dataset_id": details.task.result_payload.get("result_dataset_id") if details.task.result_payload else None,
                "artifact_id": generic_artifact.id,
                "row_count": details.task.result_payload.get("row_count") if details.task.result_payload else None,
            },
            content_blocks=[{"type": "markdown", "text": f"Apply results are ready: {link}"}],
        )

    def _resolve_apply_input_sources(self, input_sources: list[str]) -> list[str]:
        return [self._resolve_apply_input_source(input_source) for input_source in input_sources]

    def _resolve_apply_input_source(self, input_source: str) -> str:
        source = input_source.strip()
        if source.startswith("artifact://"):
            artifact = self._artifact_service.resolve_uri(source)
            if not artifact.exists:
                raise ValidationError("Apply input artifact file is missing.")
            return artifact.absolute_path

        try:
            dataset = self._dataset_service.get_dataset(source)
        except NotFoundError:
            raise ValidationError("model.apply input_sources must be registered dataset ids or artifact:// URIs.") from None
        if not Path(dataset.source_path).exists():
            raise ValidationError("Apply input dataset source file is missing.")
        return dataset.source_path

    def _model_task_query(self, arguments: dict[str, Any], context: ToolExecutionContext) -> ToolExecutionResult:
        self._raise_if_cancelled(context)
        task_ids = self._require_string_list(arguments, "task_ids")
        include_logs = bool(arguments.get("include_logs"))
        try:
            max_log_entries = int(
                arguments.get("max_log_entries") if arguments.get("max_log_entries") is not None else 200
            )
        except (TypeError, ValueError) as exc:
            raise ValidationError("max_log_entries must be an integer.") from exc
        if max_log_entries < 0:
            raise ValidationError("max_log_entries must be greater than or equal to 0.")
        if max_log_entries > 1000:
            raise ValidationError("max_log_entries must be less than or equal to 1000.")

        tasks = [
            self._ml_task_details_payload(
                task_id,
                include_logs=include_logs,
                max_log_entries=max_log_entries,
            )
            for task_id in task_ids
        ]
        payload = {
            "task_ids": task_ids,
            "tasks": tasks,
        }
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[{"type": "markdown", "text": self._model_task_query_markdown(tasks)}],
        )

    def _resolve_sql_bindings(self, arguments: dict[str, Any], *, tool_name: str) -> list[DatasetSqlBinding]:
        raw_bindings = arguments.get("bindings")
        raw_dataset_id = str(arguments.get("dataset_id") or "").strip()
        if raw_bindings:
            if raw_dataset_id:
                raise ValidationError(f"{tool_name} accepts dataset_id for one input or bindings for multiple inputs.")
            if not isinstance(raw_bindings, list):
                raise ValidationError(f"{tool_name} bindings must be a list.")
            bindings: list[DatasetSqlBinding] = []
            for raw_binding in raw_bindings:
                if not isinstance(raw_binding, dict):
                    raise ValidationError(f"{tool_name} bindings must contain objects.")
                alias = str(raw_binding.get("alias") or "").strip()
                dataset_id = str(raw_binding.get("dataset_id") or "").strip()
                if not alias or not dataset_id:
                    raise ValidationError(f"{tool_name} bindings require alias and dataset_id.")
                dataset = self._dataset_service.get_dataset(dataset_id)
                bindings.append(
                    DatasetSqlBinding(
                        alias=alias,
                        dataset_id=dataset.id,
                        source_path=dataset.source_path,
                    )
                )
            return bindings

        dataset_id = self._require_string(arguments, "dataset_id")
        dataset = self._dataset_service.get_dataset(dataset_id)
        return [
            DatasetSqlBinding(
                alias="input",
                dataset_id=dataset.id,
                source_path=dataset.source_path,
            )
        ]

    def _query_result_markdown(self, payload: dict[str, Any]) -> str:
        returned = int(payload.get("returned_row_count") or 0)
        limit = int(payload.get("limit") or 0)
        suffix = " (truncated)" if payload.get("truncated") else ""
        lines = [f"Query returned {returned} row(s) with limit {limit}{suffix}."]
        rows = payload.get("rows")
        columns = payload.get("columns")
        if not isinstance(rows, list) or not rows:
            return "\n".join(lines)
        if not isinstance(columns, list) or not columns:
            return "\n".join(lines)
        column_names = [str(column.get("name")) for column in columns if isinstance(column, dict)]
        preview_rows = rows[:10]
        lines.extend(
            [
                "",
                "| " + " | ".join(self._markdown_cell(column) for column in column_names) + " |",
                "| " + " | ".join("---" for _column in column_names) + " |",
            ]
        )
        for row in preview_rows:
            if not isinstance(row, dict):
                continue
            lines.append(
                "| "
                + " | ".join(self._markdown_cell(row.get(column)) for column in column_names)
                + " |"
            )
        return "\n".join(lines)

    def _markdown_cell(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("|", "\\|")

    def _register_generated_dataset_result(
        self,
        context: ToolExecutionContext,
        *,
        output_path: Path,
        name: str,
        summary: str,
        derived_from_dataset_id: str | None = None,
        metadata_payload: dict[str, Any] | None = None,
    ) -> ToolExecutionResult:
        dataset = self._dataset_service.register_dataset(
            RegisterDatasetInput(
                source_path=str(output_path.resolve()),
                name=name,
                derived_from_dataset_id=derived_from_dataset_id,
            )
        )
        inspection = self._dataset_service.inspect_source_file(InspectDatasetInput(source_path=dataset.source_path))
        inspection_payload = self._inspection_payload(inspection)
        artifact_metadata = dict(metadata_payload or {})
        if derived_from_dataset_id:
            artifact_metadata["derived_from_dataset_id"] = derived_from_dataset_id
        artifact = self._register_dataset_artifact(
            context,
            title=dataset.name,
            path=Path(dataset.source_path),
            dataset_id=dataset.id,
            preview_payload=inspection_payload,
            metadata_payload=artifact_metadata,
        )
        link = build_artifact_markdown_link(artifact)
        return ToolExecutionResult(
            payload={
                "dataset_id": dataset.id,
                "artifact_id": artifact.id,
                "inspection": inspection_payload,
            },
            content_blocks=[{"type": "markdown", "text": f"{summary} {link}"}],
        )

    def _inspection_payload(self, inspection: DatasetInspection) -> dict[str, Any]:
        return inspection.model_dump(mode="json", exclude={"source_path"})

    def _structure_payload(self, source_path: Path, inspection: DatasetInspection) -> dict[str, Any]:
        schema = resolve_tabular_schema(inspection.preview_columns)
        row_windows, declared_dimensions = self._structure_row_windows(source_path, inspection.source_format)
        layout_evidence = self._layout_evidence(row_windows, declared_dimensions)
        return {
            "format": inspection.source_format.value,
            "sheet": self._structure_sheet(source_path, inspection.source_format),
            "coordinate_system": {
                "rows": (
                    "spreadsheet_1_based"
                    if inspection.source_format in {DatasetSourceFormat.XLSX, DatasetSourceFormat.XLS}
                    else "file_1_based"
                ),
                "columns": "position_0_based",
            },
            "declared_dimensions": declared_dimensions,
            "observed_dimensions": {
                "rows": inspection.row_count + 1,
                "columns": inspection.column_count,
            },
            "layout_evidence": layout_evidence,
            "row_windows": row_windows,
            "columns": self._structure_columns(schema, row_windows),
        }

    def _structure_row_windows(
        self,
        source_path: Path,
        source_format: DatasetSourceFormat,
        *,
        max_rows: int = 5,
        max_columns: int = 50,
    ) -> tuple[list[dict[str, Any]], dict[str, int | None] | None]:
        if source_format is DatasetSourceFormat.CSV:
            return self._csv_structure_row_windows(source_path, max_rows=max_rows, max_columns=max_columns), None
        if source_format in {DatasetSourceFormat.XLSX, DatasetSourceFormat.XLS}:
            return self._xlsx_structure_row_windows(source_path, max_rows=max_rows, max_columns=max_columns)
        return [], None

    def _csv_structure_row_windows(
        self,
        source_path: Path,
        *,
        max_rows: int,
        max_columns: int,
    ) -> list[dict[str, Any]]:
        windows: list[dict[str, Any]] = []
        with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row_number, row in enumerate(reader, start=1):
                if row_number > max_rows:
                    break
                windows.append(self._row_window_payload(row_number, row[:max_columns]))
        return windows

    def _xlsx_structure_row_windows(
        self,
        source_path: Path,
        *,
        max_rows: int,
        max_columns: int,
    ) -> tuple[list[dict[str, Any]], dict[str, int | None]]:
        workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            declared_dimensions = {
                "rows": int(worksheet.max_row or 0) if worksheet.max_row is not None else None,
                "columns": int(worksheet.max_column or 0) if worksheet.max_column is not None else None,
            }
            reset_dimensions = getattr(worksheet, "reset_dimensions", None)
            if callable(reset_dimensions):
                reset_dimensions()
            windows = []
            for row_number, row in enumerate(
                worksheet.iter_rows(max_row=max_rows, max_col=max_columns, values_only=True),
                start=1,
            ):
                windows.append(self._row_window_payload(row_number, list(row)))
            return windows, declared_dimensions
        finally:
            workbook.close()

    def _structure_sheet(self, source_path: Path, source_format: DatasetSourceFormat) -> dict[str, Any] | None:
        if source_format not in {DatasetSourceFormat.XLSX, DatasetSourceFormat.XLS}:
            return None
        workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            return {"index": 0, "name": workbook.active.title}
        finally:
            workbook.close()

    def _row_window_payload(self, row_number: int, cells: list[Any]) -> dict[str, Any]:
        formatted = [self._structure_cell(value) for value in cells]
        observed_width = 0
        for index in range(len(formatted) - 1, -1, -1):
            if formatted[index].strip():
                observed_width = index + 1
                break
        return {
            "row": row_number,
            "non_empty_count": sum(1 for value in formatted if value.strip()),
            "observed_width": observed_width,
            "cells": formatted[:observed_width or 1],
        }

    def _structure_cell(self, value: Any, *, max_length: int = 160) -> str:
        if value is None:
            return ""
        text = str(value).replace("\r", " ").replace("\n", " ").strip()
        if len(text) <= max_length:
            return text
        return text[: max_length - 3] + "..."

    def _layout_evidence(
        self,
        row_windows: list[dict[str, Any]],
        declared_dimensions: dict[str, int | None] | None,
    ) -> list[str]:
        evidence: list[str] = []
        if declared_dimensions and (
            (declared_dimensions.get("rows") or 0) <= 1
            or (declared_dimensions.get("columns") or 0) <= 1
        ):
            evidence.append("declared_dimension_mismatch")
        widths = [int(row.get("observed_width") or 0) for row in row_windows]
        if len(widths) >= 2 and widths[0] <= 1:
            evidence.append("leading_sparse_rows")
        if widths and max(widths) > max(widths[0], 1):
            evidence.append("dense_rows_after_sparse_rows")
        return evidence

    def _structure_columns(self, schema: TabularSchema, row_windows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        payload = tabular_schema_payload(schema)
        columns: list[dict[str, Any]] = []
        for column in payload["columns"]:
            index = int(column["index"])
            samples = []
            for row in row_windows:
                cells = row.get("cells")
                if isinstance(cells, list) and index < len(cells):
                    samples.append(cells[index])
            columns.append({**column, "sample_values": samples[:5]})
        return columns

    def _data_peek_provider_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        provider_payload = {
            "dataset_id": payload.get("dataset_id"),
            "inspection": payload.get("inspection"),
            "structure": payload.get("structure"),
        }
        analysis = payload.get("analysis")
        if isinstance(analysis, dict):
            provider_payload["analysis"] = self._compact_analysis_payload(analysis)
        return provider_payload

    def _compact_analysis_payload(self, analysis: dict[str, Any]) -> dict[str, Any]:
        if not analysis.get("enabled"):
            return {"enabled": False}
        profile = analysis.get("profile")
        if not isinstance(profile, dict):
            return {"enabled": True}
        compact = {
            "enabled": True,
            "basic_info": profile.get("basic_info"),
            "field_type_summary": profile.get("field_type_summary"),
        }
        limits = profile.get("limits")
        if isinstance(limits, dict):
            compact["limits"] = limits
        return compact

    def _register_dataset_artifact(
        self,
        context: ToolExecutionContext,
        *,
        title: str,
        path: Path,
        dataset_id: str,
        preview_payload: dict[str, Any],
        metadata_payload: dict[str, Any] | None = None,
    ):
        return self._artifact_service.register_artifact(
            RegisterArtifactInput(
                thread_id=context.thread_id,
                turn_id=context.turn_id,
                tool_call_id=context.tool_call_id,
                kind=ArtifactKind.DATASET,
                title=title,
                absolute_path=str(path.resolve()),
                mime_type="text/csv" if path.suffix.lower() == ".csv" else None,
                preview_payload=preview_payload,
                metadata_payload={"dataset_id": dataset_id, **(metadata_payload or {})},
            )
        )

    def _lambda_artifact_kind(self, raw_kind: str) -> ArtifactKind:
        value = str(raw_kind or "").strip()
        if value == "image":
            return ArtifactKind.IMAGE
        if value == "report":
            return ArtifactKind.REPORT
        if value in {"dataset", "file", "table"}:
            return ArtifactKind.FILE
        try:
            return ArtifactKind(value)
        except ValueError:
            return ArtifactKind.OTHER

    def _rewrite_lambda_artifact_uris(self, value: Any, artifact_map: dict[str, str]) -> Any:
        if isinstance(value, str):
            rewritten = value
            for placeholder_id, artifact_id in artifact_map.items():
                rewritten = rewritten.replace(f"artifact://{placeholder_id}", build_artifact_uri(artifact_id))
            return rewritten
        if isinstance(value, list):
            return [self._rewrite_lambda_artifact_uris(item, artifact_map) for item in value]
        if isinstance(value, dict):
            return {
                key: self._rewrite_lambda_artifact_uris(item, artifact_map)
                for key, item in value.items()
            }
        return value

    def _load_frame(self, path: Path) -> pd.DataFrame:
        source_format = detect_source_format(path)
        if source_format.value == "unknown":
            raise ValidationError("Only .csv, .xlsx, and .xls dataset files are supported.")
        return load_dataframe(path, source_format)

    def _wait_for_training_models_or_none(
        self,
        root_task_ids: list[str],
        *,
        context: ToolExecutionContext,
        timeout_seconds: float,
    ) -> tuple[list[MLTaskRow], list[TrainedModelRow]] | None:
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            root_tasks = [self._ml_service.get_task_details(task_id).task for task_id in root_task_ids]
            trained_models = self._trained_models_for_root_tasks(root_task_ids)
            related_tasks = self._related_training_tasks(root_tasks, trained_models)
            self._raise_if_cancelled(
                context,
                ml_task_ids=[task.id for task in related_tasks] or root_task_ids,
            )

            failed = [
                task
                for task in related_tasks
                if task.status in {MLTaskStatus.FAILED, MLTaskStatus.CANCELLED}
            ]
            if failed:
                raise ValidationError(f"ML task '{failed[0].id}' finished with status '{failed[0].status.value}'.")

            root_tasks_succeeded = all(task.status is MLTaskStatus.SUCCEEDED for task in root_tasks)
            if root_tasks_succeeded and len(trained_models) == len(root_task_ids):
                pending_evaluation = False
                models_by_root_task = {model.ml_task_id: model for model in trained_models}
                for root_task in root_tasks:
                    model = models_by_root_task.get(root_task.id)
                    if model is None:
                        pending_evaluation = True
                        break
                    if self._training_task_requires_follow_up_evaluation(root_task):
                        evaluation_task_id = self._evaluation_task_id_for_model(model)
                        if not evaluation_task_id:
                            pending_evaluation = True
                            break
                        evaluation_task = self._ml_service.get_task_details(evaluation_task_id).task
                        if evaluation_task.status is not MLTaskStatus.SUCCEEDED:
                            pending_evaluation = True
                            break
                if not pending_evaluation:
                    return self._related_training_tasks(root_tasks, trained_models), trained_models

            time.sleep(0.1)
        return None

    def _wait_for_task(self, task_id: str, *, context: ToolExecutionContext, timeout_seconds: float = 120.0):
        task = self._wait_for_task_or_none(task_id, context=context, timeout_seconds=timeout_seconds)
        if task is None:
            raise ValidationError(f"Timed out waiting for ML task '{task_id}'.")
        return task

    def _wait_for_task_or_none(
        self,
        task_id: str,
        *,
        context: ToolExecutionContext,
        timeout_seconds: float,
    ) -> MLTaskRow | None:
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            self._raise_if_cancelled(context, ml_task_ids=[task_id])
            task = self._ml_service.get_task_details(task_id).task
            if task.status in self._terminal_statuses():
                if task.status is not MLTaskStatus.SUCCEEDED:
                    raise ValidationError(f"ML task '{task.id}' finished with status '{task.status.value}'.")
                return task
            time.sleep(0.1)
        return None

    def _raise_if_cancelled(self, context: ToolExecutionContext, *, ml_task_ids: list[str] | None = None) -> None:
        if not context.cancel_requested():
            return
        if ml_task_ids:
            for task_id in ml_task_ids:
                try:
                    self._ml_service.cancel_task(task_id)
                except Exception:
                    continue
        raise ValidationError("Agent run was cancelled.")

    def _terminal_statuses(self) -> set[MLTaskStatus]:
        return {MLTaskStatus.SUCCEEDED, MLTaskStatus.FAILED, MLTaskStatus.CANCELLED}

    def _trained_models_for_root_tasks(self, root_task_ids: list[str]) -> list[TrainedModelRow]:
        models_by_task_id: dict[str, TrainedModelRow] = {}
        for task_id in root_task_ids:
            model = self._ml_service.get_trained_model_by_ml_task(task_id)
            if model is not None:
                models_by_task_id[task_id] = model
        return [models_by_task_id[task_id] for task_id in root_task_ids if task_id in models_by_task_id]

    def _related_training_tasks(
        self,
        root_tasks: list[MLTaskRow],
        trained_models: list[TrainedModelRow],
    ) -> list[MLTaskRow]:
        tasks: list[MLTaskRow] = []
        seen_task_ids: set[str] = set()
        for task in root_tasks:
            tasks.append(task)
            seen_task_ids.add(task.id)
        for model in trained_models:
            evaluation_task_id = self._evaluation_task_id_for_model(model)
            if not evaluation_task_id or evaluation_task_id in seen_task_ids:
                continue
            task = self._ml_service.get_task_details(evaluation_task_id).task
            tasks.append(task)
            seen_task_ids.add(task.id)
        return tasks

    def _evaluation_task_id_for_model(self, model: TrainedModelRow) -> str | None:
        task_id = model.metadata_payload.get("evaluation_ml_task_id")
        if isinstance(task_id, str) and task_id.strip():
            return task_id
        return None

    def _training_task_requires_follow_up_evaluation(self, task: MLTaskRow) -> bool:
        continuation = task.request_payload.get("continuation_plan")
        return isinstance(continuation, dict) and continuation.get("next_operation") == "evaluate"

    def _training_task_receipt(
        self,
        *,
        tool_name: str,
        dataset_id: str,
        root_task_ids: list[str],
        operation: str,
    ) -> ToolExecutionResult:
        root_tasks = [self._ml_service.get_task_details(task_id).task for task_id in root_task_ids]
        trained_models = self._trained_models_for_root_tasks(root_task_ids)
        tasks = self._related_training_tasks(root_tasks, trained_models)
        task_ids = [task.id for task in tasks] or list(root_task_ids)
        payload = {
            "async_state": "running_background",
            "dataset_id": dataset_id,
            "operation": operation,
            "task_ids": task_ids,
            "root_task_ids": list(root_task_ids),
            "ml_tasks": [self._ml_task_payload(task) for task in tasks],
            "trained_models": [self._trained_model_payload(model) for model in trained_models],
        }
        summary = (
            "Model tuning running in background"
            if tool_name == "model.hyper_train"
            else "Model training running in background"
        )
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[
                {"type": "tool_event_summary", "text": summary},
                {"type": "markdown", "text": self._task_receipt_markdown(payload)},
            ],
        )

    def _single_task_receipt(
        self,
        *,
        tool_name: str,
        task_id: str,
        operation: str,
    ) -> ToolExecutionResult:
        task = self._ml_service.get_task_details(task_id).task
        payload = {
            "async_state": "running_background",
            "operation": operation,
            "ml_task_id": task.id,
            "task_ids": [task.id],
            "root_task_ids": [task.id],
            "dataset_id": task.dataset_id,
            "ml_tasks": [self._ml_task_payload(task)],
        }
        summary = "Model apply running in background" if tool_name == "model.apply" else "ML task running in background"
        return ToolExecutionResult(
            payload=payload,
            content_blocks=[
                {"type": "tool_event_summary", "text": summary},
                {"type": "markdown", "text": self._task_receipt_markdown(payload)},
            ],
        )

    def _task_receipt_markdown(self, payload: dict[str, Any]) -> str:
        task_ids = [str(task_id) for task_id in payload.get("task_ids", [])]
        lines = ["The ML work is still running in the background."]
        if task_ids:
            lines.append("")
            lines.append("Task ids:")
            lines.extend(f"- `{task_id}`" for task_id in task_ids)
        lines.append("")
        lines.append("Use `model.task.query` with these task ids to inspect status and logs.")
        return "\n".join(lines)

    def _ml_task_payload(self, task: MLTaskRow) -> dict[str, Any]:
        return {
            "task_id": task.id,
            "dataset_id": task.dataset_id,
            "task_type": task.task_type.value,
            "status": task.status.value,
            "model_key": self._model_key_from_task_payload(task.request_payload),
            "error_summary": task.error_summary,
            "created_at": task.created_at.isoformat() if task.created_at else None,
            "started_at": task.started_at.isoformat() if task.started_at else None,
            "finished_at": task.finished_at.isoformat() if task.finished_at else None,
            "updated_at": task.updated_at.isoformat() if task.updated_at else None,
            "follow_up_task_ids": self._follow_up_task_ids(task),
        }

    def _ml_task_details_payload(
        self,
        task_id: str,
        *,
        include_logs: bool,
        max_log_entries: int,
    ) -> dict[str, Any]:
        details = self._ml_service.get_task_details(task_id)
        task_payload = self._ml_task_payload(details.task)
        logs = [log.model_dump(mode="json") for log in details.logs[-max_log_entries:]] if include_logs and max_log_entries else []
        task_payload.update(
            {
                "request": self._task_request_summary(details.task.request_payload),
                "result": dict(details.task.result_payload or {}),
                "artifacts": [
                    {
                        "artifact_id": artifact.id,
                        "artifact_kind": artifact.artifact_kind.value,
                        "absolute_path": artifact.absolute_path,
                        "ready_to_open": artifact.ready_to_open,
                        "created_at": artifact.created_at.isoformat() if artifact.created_at else None,
                    }
                    for artifact in details.artifacts
                ],
                "logs": logs,
            }
        )
        return task_payload

    def _task_request_summary(self, request_payload: dict[str, Any]) -> dict[str, Any]:
        keys = [
            "project_id",
            "dataset_id",
            "evaluation_kind",
            "feature_columns",
            "manual_training",
            "hyperparameter_tuning",
            "evaluate_model",
            "apply_model",
            "input_sources",
        ]
        return {key: request_payload[key] for key in keys if key in request_payload}

    def _model_key_from_task_payload(self, request_payload: dict[str, Any]) -> str | None:
        for key in ("manual_training", "hyperparameter_tuning", "evaluate_model", "apply_model"):
            value = request_payload.get(key)
            if isinstance(value, dict) and isinstance(value.get("model_key"), str):
                return value["model_key"]
        return None

    def _follow_up_task_ids(self, task: MLTaskRow) -> list[str]:
        if task.task_type not in {MLTaskType.FIT, MLTaskType.HYPERPARAMETER_TUNING}:
            return []
        model = self._ml_service.get_trained_model_by_ml_task(task.id)
        if model is None:
            return []
        evaluation_task_id = self._evaluation_task_id_for_model(model)
        return [evaluation_task_id] if evaluation_task_id else []

    def _trained_model_payload(self, model: TrainedModelRow) -> dict[str, Any]:
        return {
            "trained_model_id": model.id,
            "dataset_id": model.dataset_id,
            "model_key": model.model_key,
            "artifact_path": model.artifact_path,
            "metadata": dict(model.metadata_payload),
        }

    def _training_summary_markdown(self, payload: dict[str, Any]) -> str:
        models = payload.get("trained_models", [])
        lines = ["Training completed."]
        for model in models:
            metadata = model.get("metadata") if isinstance(model, dict) else None
            metric_text = self._metadata_evaluation_summary(metadata if isinstance(metadata, dict) else {})
            line = f"- `{model['model_key']}` trained model id: `{model['trained_model_id']}`"
            if metric_text:
                line += f"; evaluation: {metric_text}"
            lines.append(line)
        return "\n".join(lines)

    def _model_task_query_markdown(self, tasks: list[dict[str, Any]]) -> str:
        if not tasks:
            return "No ML tasks were found."
        lines = ["ML task status:"]
        for task in tasks:
            task_id = task.get("task_id")
            task_type = task.get("task_type")
            status = task.get("status")
            model_key = task.get("model_key") or "unknown model"
            lines.append(f"- `{task_id}` {task_type} for `{model_key}`: `{status}`")
            error_summary = str(task.get("error_summary") or "").strip()
            if error_summary:
                lines.append(f"  Error: {error_summary}")
            evaluation_lines = self._task_evaluation_summary_lines(task)
            lines.extend(evaluation_lines)
            follow_up_task_ids = task.get("follow_up_task_ids")
            if isinstance(follow_up_task_ids, list) and follow_up_task_ids:
                joined_ids = ", ".join(f"`{task_id}`" for task_id in follow_up_task_ids)
                lines.append(f"  Follow-up task ids: {joined_ids}")
        return "\n".join(lines)

    def _metadata_evaluation_summary(self, metadata: dict[str, Any]) -> str:
        metric_name = metadata.get("evaluation_primary_metric_name")
        metric_value = metadata.get("evaluation_primary_metric_value")
        if not isinstance(metric_name, str) or metric_value is None:
            return ""
        return f"{metric_name}={self._format_metric_value(metric_value)}"

    def _task_evaluation_summary_lines(self, task: dict[str, Any]) -> list[str]:
        if task.get("task_type") != MLTaskType.EVALUATE.value:
            return []
        result = task.get("result")
        if not isinstance(result, dict):
            return []
        evaluation = result.get("evaluation")
        if not isinstance(evaluation, dict):
            return []
        metrics = evaluation.get("metrics")
        if not isinstance(metrics, dict):
            return []

        lines: list[str] = []
        primary_name = evaluation.get("primary_metric_name")
        primary_value = evaluation.get("primary_metric_value")
        if isinstance(primary_name, str) and primary_value is not None:
            lines.append(f"  Primary metric: {primary_name}={self._format_metric_value(primary_value)}")

        metric_names = self._summary_metric_names(str(result.get("evaluation_kind") or ""), metrics)
        metric_text = self._format_metric_list(metrics, metric_names)
        if metric_text:
            lines.append(f"  Key metrics: {metric_text}")

        details = evaluation.get("details")
        if isinstance(details, dict):
            probability_metrics = details.get("probability_metrics")
            if isinstance(probability_metrics, dict) and probability_metrics.get("available") is False:
                reason = probability_metrics.get("reason")
                if isinstance(reason, str) and reason:
                    lines.append(f"  Probability metrics unavailable: {reason}.")
        return lines

    def _summary_metric_names(self, evaluation_kind: str, metrics: dict[str, Any]) -> list[str]:
        if evaluation_kind == EvaluationKind.REGRESSION.value:
            return ["r2", "rmse", "mae", "mape", "explained_variance"]
        if evaluation_kind == EvaluationKind.CLASSIFICATION.value:
            names = [
                "accuracy",
                "balanced_accuracy",
                "f1_macro",
                "f1_weighted",
                "roc_auc",
                "pr_auc",
                "log_loss",
            ]
            return [name for name in names if name in metrics]
        return list(metrics)[:5]

    def _format_metric_list(self, metrics: dict[str, Any], names: list[str]) -> str:
        parts: list[str] = []
        for name in names:
            if name not in metrics:
                continue
            parts.append(f"{name}={self._format_metric_value(metrics[name])}")
        return ", ".join(parts)

    def _format_metric_value(self, value: Any) -> str:
        if isinstance(value, (int, float)):
            return f"{float(value):.4g}"
        return str(value)

    def _model_catalog_payload(
        self,
        entry,
        *,
        detail_query: bool,
        include_param_schema: bool,
        include_param_grid_schema: bool,
    ) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "model_key": entry.model_key,
            "display_name": entry.display_name,
            "description": entry.guidance,
            "problem_kind": entry.problem_kind.value if entry.problem_kind is not None else None,
            "evaluation_kind": entry.evaluation_kind.value,
            "model_family": entry.model_family.value,
            "model_task_kind": entry.model_task_kind.value,
            "family": entry.family,
            "recommendation_tier": entry.recommendation_tier,
            "supports_fit": entry.supports_fit,
            "supports_hyperparameter_tuning": entry.supports_hyperparameter_tuning,
        }
        if detail_query:
            payload.update(
                {
                    "summary_metric_name": entry.summary_metric_name,
                    "requires_target": entry.requires_target,
                    "train_role_schema": entry.train_role_schema.model_dump(mode="json"),
                    "apply_role_schema": entry.apply_role_schema.model_dump(mode="json"),
                    "result_contract": entry.result_contract.model_dump(mode="json"),
                }
            )
        if include_param_schema:
            payload["param_schema"] = entry.param_schema
        if include_param_grid_schema:
            payload["param_grid_schema"] = entry.param_grid_schema
        return payload

    def _model_metadata_markdown(self, models: list[dict[str, Any]]) -> str:
        if not models:
            return "No models match the requested filters."
        lines = ["Available models:"]
        for model in models:
            capabilities = ["fit"] if model["supports_fit"] else []
            if model["supports_hyperparameter_tuning"]:
                capabilities.append("hyperparameter_tuning")
            legacy_problem = f", {model['problem_kind']}" if model["problem_kind"] else ""
            lines.append(
                "- "
                f"`{model['model_key']}` ({model['display_name']}{legacy_problem}); "
                f"evaluation: {model['evaluation_kind']}; "
                f"task: {model['model_task_kind']}; "
                f"capabilities: {', '.join(capabilities)}"
            )
        return "\n".join(lines)

    def _normalize_model_mapping(
        self,
        raw_mapping: Any,
        *,
        field_name: str,
        require_hyperparameter_tuning: bool = False,
    ) -> dict[str, Any]:
        if raw_mapping is None:
            return {}
        if not isinstance(raw_mapping, dict):
            raise ValidationError(f"{field_name} must be an object keyed by model key.")
        normalized: dict[str, Any] = {}
        failures: list[str] = []
        for raw_key, value in raw_mapping.items():
            model_key = self._canonical_model_key(str(raw_key))
            if model_key is None:
                failures.append(str(raw_key))
                continue
            if value is None:
                value = {}
            if not isinstance(value, dict):
                failures.append(f"{raw_key} must map to an object")
                continue
            if require_hyperparameter_tuning:
                catalog = get_model_catalog_entry(model_key)
                if not catalog.supports_hyperparameter_tuning:
                    failures.append(f"{raw_key} lacks hyperparameter_tuning support")
                    continue
            normalized[model_key] = value
        if failures:
            raise ValidationError(self._model_key_error_message(field_name, failures))
        return normalized

    def _normalize_model_keys(
        self,
        raw_keys: list[str],
        *,
        field_name: str,
        require_hyperparameter_tuning: bool = False,
    ) -> list[str]:
        normalized: list[str] = []
        failures: list[str] = []
        for raw_key in raw_keys:
            model_key = self._canonical_model_key(raw_key)
            if model_key is None:
                failures.append(raw_key)
                continue
            if require_hyperparameter_tuning:
                catalog = get_model_catalog_entry(model_key)
                if not catalog.supports_hyperparameter_tuning:
                    failures.append(f"{raw_key} lacks hyperparameter_tuning support")
                    continue
            if model_key not in normalized:
                normalized.append(model_key)
        if failures:
            raise ValidationError(self._model_key_error_message(field_name, failures))
        return normalized

    def _canonical_model_key(self, raw_key: str) -> str | None:
        value = raw_key.strip()
        available = set(list_model_keys())
        if value in available:
            return value
        lowered = value.lower()
        if lowered in available:
            return lowered
        for token in self._model_key_alias_tokens(value):
            aliased = self._model_key_aliases.get(token)
            if aliased in available:
                return aliased
        return None

    def _build_model_key_aliases(self) -> dict[str, str]:
        aliases: dict[str, str] = {}
        priorities: dict[str, int] = {}
        for entry in list_model_catalog():
            priority = self._model_alias_priority(entry)
            for token in self._model_entry_alias_tokens(entry):
                if priority < priorities.get(token, 100):
                    aliases[token] = entry.model_key
                    priorities[token] = priority
        aliases.update(_MODEL_KEY_ALIAS_OVERRIDES)
        return aliases

    def _model_entry_alias_tokens(self, entry) -> set[str]:
        leaf_key = entry.model_key.split(".", 1)[-1]
        values = {
            entry.model_key,
            entry.model_key.replace(".", "_"),
            leaf_key,
            entry.display_name,
            f"{entry.evaluation_kind.value}_{leaf_key}",
            f"{entry.model_family.value}_{leaf_key}",
            f"{entry.model_task_kind.value}_{leaf_key}",
        }
        if entry.problem_kind is not None:
            values.add(f"{entry.problem_kind.value}_{leaf_key}")
        tokens: set[str] = set()
        for value in values:
            for token in self._model_key_alias_tokens(value):
                tokens.add(token)
                stripped = self._strip_model_alias_suffix(token)
                if stripped:
                    tokens.update(self._model_key_alias_tokens(stripped))
        return tokens

    def _model_alias_priority(self, entry) -> int:
        if entry.evaluation_kind is EvaluationKind.REGRESSION:
            return 0
        if entry.evaluation_kind is EvaluationKind.CLASSIFICATION:
            return 1
        task_order = {
            ModelTaskKind.SEGMENTER: 2,
            ModelTaskKind.ANOMALY_SCORER: 3,
            ModelTaskKind.RULE_MINER: 4,
            ModelTaskKind.RECOMMENDER: 5,
        }
        return task_order.get(entry.model_task_kind, 100)

    def _model_key_alias_tokens(self, value: str) -> list[str]:
        token = "".join(char.lower() if char.isalnum() else "_" for char in value).strip("_")
        while "__" in token:
            token = token.replace("__", "_")
        if not token:
            return []
        compact = token.replace("_", "")
        return [token] if compact == token else [token, compact]

    def _strip_model_alias_suffix(self, token: str) -> str:
        parts = token.split("_")
        if len(parts) > 1 and parts[-1] in _MODEL_ALIAS_SUFFIXES:
            return "_".join(parts[:-1])
        return ""

    def _model_key_error_message(self, field_name: str, failures: list[str]) -> str:
        return (
            f"{field_name} contains unsupported model keys: {', '.join(failures)}. "
            "Call model.metadata to inspect available canonical model keys. "
            f"Available keys: {', '.join(list_model_keys())}."
        )

    def _require_string(self, arguments: dict[str, Any], key: str) -> str:
        value = str(arguments.get(key) or "").strip()
        if not value:
            raise ValidationError(f"{key} is required.")
        return value

    def _require_string_list(self, arguments: dict[str, Any], key: str) -> list[str]:
        values = arguments.get(key)
        if not isinstance(values, list):
            raise ValidationError(f"{key} must be a list.")
        normalized = [str(value).strip() for value in values if str(value).strip()]
        if not normalized:
            raise ValidationError(f"{key} cannot be empty.")
        return normalized

    def _optional_string_list(self, arguments: dict[str, Any], key: str) -> list[str]:
        values = arguments.get(key)
        if values is None:
            return []
        if not isinstance(values, list):
            raise ValidationError(f"{key} must be a list.")
        return [str(value).strip() for value in values if str(value).strip()]

    def _optional_integer(self, arguments: dict[str, Any], key: str, *, default: int) -> int:
        value = arguments.get(key)
        if value is None:
            return default
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"{key} must be an integer.") from exc

    def _optional_boolean(self, arguments: dict[str, Any], key: str, *, default: bool) -> bool:
        value = arguments.get(key)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        raise ValidationError(f"{key} must be a boolean.")

    def _slug(self, value: str) -> str:
        normalized = "".join(char.lower() if char.isalnum() else "-" for char in value).strip("-")
        return normalized or "dataset"
