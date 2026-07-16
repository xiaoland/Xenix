from __future__ import annotations

import csv
from enum import StrEnum
from pathlib import Path

import pandas as pd
import polars as pl
import duckdb
from pydantic import BaseModel, Field

from ..exceptions import ValidationError
from .storage.models import DatasetSourceFormat
from .tabular import (
    apply_tabular_schema,
    load_tabular_frame,
    load_tabular_schema,
    preview_rows,
    reconcile_tabular_schema_to_loaded_columns,
    resolve_tabular_schema_for_loaded_frame,
    tabular_schema_tool_names,
)


_XLSX_ATTACHMENT_SCAN_ROW_LIMIT = 50
_XLSX_ROW_MARKER = b"<row "


class DatasetColumnKind(StrEnum):
    NUMERIC = "numeric"
    BOOLEAN = "boolean"
    CATEGORICAL = "categorical"
    DATETIME = "datetime"
    TEXT = "text"
    UNKNOWN = "unknown"


class InspectDatasetInput(BaseModel):
    source_path: str


class DatasetColumnMetadata(BaseModel):
    name: str
    kind: DatasetColumnKind
    nullable: bool


class DatasetInspection(BaseModel):
    source_path: str
    source_format: DatasetSourceFormat
    file_name: str
    row_count: int
    column_count: int
    columns: list[DatasetColumnMetadata]
    preview_columns: list[str] = Field(default_factory=list)
    preview_rows: list[list[str]] = Field(default_factory=list)


class DatasetAttachmentMetadata(BaseModel):
    source_path: str
    source_format: DatasetSourceFormat
    file_name: str
    row_count: int
    column_count: int
    preview_columns: list[str] = Field(default_factory=list)


def detect_source_format(path: Path) -> DatasetSourceFormat:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return DatasetSourceFormat.CSV
    if suffix == ".parquet":
        return DatasetSourceFormat.PARQUET
    if suffix == ".xlsx":
        return DatasetSourceFormat.XLSX
    if suffix == ".xls":
        return DatasetSourceFormat.XLS
    return DatasetSourceFormat.UNKNOWN


def load_dataframe(path: Path, source_format: DatasetSourceFormat) -> pd.DataFrame:
    if source_format is DatasetSourceFormat.CSV:
        return pd.read_csv(path)
    if source_format is DatasetSourceFormat.PARQUET:
        return duckdb.connect(database=":memory:").execute(
            "SELECT * FROM read_parquet(?)",
            [str(path)],
        ).fetchdf()
    if source_format in {DatasetSourceFormat.XLSX, DatasetSourceFormat.XLS}:
        return pd.read_excel(path)
    raise ValidationError("Only .csv, .parquet, .xlsx, and .xls dataset files are supported.")


def infer_column_kind(series: pl.Series) -> DatasetColumnKind:
    dtype = series.dtype
    if dtype == pl.Boolean:
        return DatasetColumnKind.BOOLEAN
    if dtype.is_numeric():
        return DatasetColumnKind.NUMERIC
    if dtype.is_temporal():
        return DatasetColumnKind.DATETIME
    if dtype in {pl.String, pl.Categorical, pl.Enum, pl.Object}:
        unique_values = series.drop_nulls().unique().head(17).to_list()
        if unique_values and len(unique_values) <= 16:
            return DatasetColumnKind.CATEGORICAL
        return DatasetColumnKind.TEXT
    return DatasetColumnKind.UNKNOWN


def inspect_dataset_file(source_path: Path) -> DatasetInspection:
    source_format = detect_source_format(source_path)
    if source_format is DatasetSourceFormat.UNKNOWN:
        raise ValidationError("Only .csv, .parquet, .xlsx, and .xls dataset files are supported.")

    dataframe = load_tabular_frame(source_path, source_format)
    if dataframe.width == 0:
        raise ValidationError("Dataset file must contain at least one column.")
    if dataframe.height == 0:
        raise ValidationError("Dataset file must contain at least one data row.")

    # Inspection is the source-schema boundary.  Keep its ordered names in
    # lockstep with the canonical resolver used by imports, SQL, and ML role
    # bindings; loader placeholders and duplicate suffixes must not leak out.
    schema = resolve_tabular_schema_for_loaded_frame(source_path, source_format, dataframe)
    canonical_names = tabular_schema_tool_names(schema)
    if len(dataframe.columns) != len(canonical_names):
        raise ValidationError("Dataset source schema could not be resolved consistently.")
    dataframe = apply_tabular_schema(dataframe, schema)

    columns = [
        DatasetColumnMetadata(
            name=str(column_name),
            kind=infer_column_kind(dataframe[column_name]),
            nullable=_has_missing_values(dataframe[column_name]),
        )
        for column_name in dataframe.columns
    ]
    preview_columns = [str(column_name) for column_name in dataframe.columns]
    return DatasetInspection(
        source_path=str(source_path),
        source_format=source_format,
        file_name=source_path.name,
        row_count=int(dataframe.height),
        column_count=int(dataframe.width),
        columns=columns,
        preview_columns=preview_columns,
        preview_rows=preview_rows(dataframe, limit=5),
    )


def inspect_attachment_metadata_file(source_path: Path) -> DatasetAttachmentMetadata:
    source_format = detect_source_format(source_path)
    if source_format is DatasetSourceFormat.UNKNOWN:
        raise ValidationError("Only .csv, .parquet, .xlsx, and .xls dataset files are supported.")
    if source_format is DatasetSourceFormat.CSV:
        return _inspect_csv_attachment_metadata(source_path, source_format)
    if source_format is DatasetSourceFormat.XLSX:
        return _inspect_xlsx_attachment_metadata(source_path, source_format)

    inspection = inspect_dataset_file(source_path)
    return DatasetAttachmentMetadata(
        source_path=inspection.source_path,
        source_format=inspection.source_format,
        file_name=inspection.file_name,
        row_count=inspection.row_count,
        column_count=inspection.column_count,
        preview_columns=inspection.preview_columns,
    )


def _inspect_csv_attachment_metadata(
    source_path: Path,
    source_format: DatasetSourceFormat,
) -> DatasetAttachmentMetadata:
    try:
        with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            try:
                header = next(reader)
            except StopIteration as exc:
                raise ValidationError("Dataset file must contain at least one column.") from exc

            header_schema = load_tabular_schema(source_path, source_format)
            schema = reconcile_tabular_schema_to_loaded_columns(header_schema, header)
            preview_columns = tabular_schema_tool_names(schema)
            if not preview_columns:
                raise ValidationError("Dataset file must contain at least one column.")

            row_count = 0
            for row in reader:
                if row and any(str(value).strip() for value in row):
                    row_count += 1
    except UnicodeDecodeError as exc:
        raise ValidationError("Unable to read dataset file.") from exc

    if row_count == 0:
        raise ValidationError("Dataset file must contain at least one data row.")
    return DatasetAttachmentMetadata(
        source_path=str(source_path),
        source_format=source_format,
        file_name=source_path.name,
        row_count=row_count,
        column_count=len(preview_columns),
        preview_columns=preview_columns,
    )


def _inspect_xlsx_attachment_metadata(
    source_path: Path,
    source_format: DatasetSourceFormat,
) -> DatasetAttachmentMetadata:
    import openpyxl

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        row_count_with_header = int(worksheet.max_row or 0)
        column_count = int(worksheet.max_column or 0)
        if row_count_with_header == 0 or column_count == 0:
            raise ValidationError("Dataset file must contain at least one column.")

        if _xlsx_declared_dimensions_are_suspicious(row_count_with_header, column_count):
            header_row, row_count, column_count = _scan_xlsx_attachment_metadata(source_path, worksheet)
        else:
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, max_col=column_count, values_only=True),
                (),
            )
            row_count = row_count_with_header - 1
    finally:
        workbook.close()

    header_schema = load_tabular_schema(source_path, source_format)
    observed_columns = [
        header_row[index] if index < len(header_row) else None
        for index in range(column_count)
    ]
    schema = reconcile_tabular_schema_to_loaded_columns(header_schema, observed_columns)
    preview_columns = tabular_schema_tool_names(schema)
    if not preview_columns:
        raise ValidationError("Dataset file must contain at least one column.")
    if row_count <= 0:
        raise ValidationError("Dataset file must contain at least one data row.")
    return DatasetAttachmentMetadata(
        source_path=str(source_path),
        source_format=source_format,
        file_name=source_path.name,
        row_count=row_count,
        column_count=column_count,
        preview_columns=preview_columns,
    )


def _xlsx_declared_dimensions_are_suspicious(row_count_with_header: int, column_count: int) -> bool:
    return row_count_with_header <= 1 or column_count <= 1


def _scan_xlsx_attachment_metadata(source_path: Path, worksheet) -> tuple[tuple[object, ...], int, int]:
    reset_dimensions = getattr(worksheet, "reset_dimensions", None)
    if callable(reset_dimensions):
        reset_dimensions()

    header_row, sampled_data_rows, column_count = _sample_xlsx_attachment_rows(worksheet)
    worksheet_path = getattr(worksheet, "_worksheet_path", "")
    row_count_with_header = _count_xlsx_worksheet_rows(source_path, worksheet_path)
    row_count = max(row_count_with_header - 1, sampled_data_rows)
    return header_row, row_count, column_count


def _sample_xlsx_attachment_rows(worksheet) -> tuple[tuple[object, ...], int, int]:
    header_row: tuple[object, ...] = ()
    sampled_data_rows = 0
    column_count = 0
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index >= _XLSX_ATTACHMENT_SCAN_ROW_LIMIT:
            break
        normalized_row = tuple(row)
        row_width = _non_empty_row_width(normalized_row)
        if row_index == 0:
            header_row = normalized_row
            column_count = max(column_count, row_width, len(normalized_row))
            continue
        if row_width:
            sampled_data_rows += 1
            column_count = max(column_count, row_width)
    return header_row, sampled_data_rows, column_count


def _count_xlsx_worksheet_rows(source_path: Path, worksheet_path: str) -> int:
    if not worksheet_path:
        return 0

    from zipfile import ZipFile

    row_count = 0
    tail = b""
    overlap_size = len(_XLSX_ROW_MARKER) - 1
    with ZipFile(source_path) as archive:
        with archive.open(worksheet_path) as worksheet_xml:
            while True:
                chunk = worksheet_xml.read(1024 * 1024)
                if not chunk:
                    break
                row_count += chunk.count(_XLSX_ROW_MARKER)
                row_count += (tail + chunk[:overlap_size]).count(_XLSX_ROW_MARKER)
                tail = chunk[-overlap_size:]
    return row_count


def _non_empty_row_width(row: tuple[object, ...]) -> int:
    for index in range(len(row) - 1, -1, -1):
        value = row[index]
        if value is not None and str(value).strip():
            return index + 1
    return 0


def _has_missing_values(series: pl.Series) -> bool:
    if series.null_count() > 0:
        return True
    if series.dtype.is_float():
        return bool(series.is_nan().sum())
    return False
